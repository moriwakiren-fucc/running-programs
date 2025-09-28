from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from PIL import Image, ImageDraw
import numpy as np
import pandas as pd

def mosaic(img_name, file_name, num_rows, num_cols):
  output_csv = "CSV_" + file_name + ".csv"
  output_img = "IMG_" + file_name + ".jpg"
  def round_to_multiple_of_8(value):
      """0〜255の範囲で、8単位に丸める"""
      rounded = int(round(value / 8) * 8)
      return min(255, max(0, rounded))

  def image_to_fixed_grid():
      """
      画像を縦num_rows×横num_colsに分割し、はみ出た部分は切り取り。
      各マスの輝度を8単位で丸めてCSV出力。
      プレビュー画像も保存。
      """
      # 画像読み込み
      img = Image.open(img_name).convert("RGB")
      width, height = img.size

      # 各マスのサイズ（切り捨て）
      cell_h = height // num_rows
      cell_w = width // num_cols

      # 実際に使う領域のサイズ
      used_w = cell_w * num_cols
      used_h = cell_h * num_rows

      # 中央基準でトリミング
      left = (width - used_w) // 2
      top = (height - used_h) // 2
      right = left + used_w
      bottom = top + used_h

      # 画像を切り取り
      img_cropped = img.crop((left, top, right, bottom))
      img_np = np.array(img_cropped)

      # 出力用キャンバス
      preview = Image.new("RGB", (used_w, used_h), (255, 255, 255))
      draw = ImageDraw.Draw(preview)

      # CSVデータ
      data = []

      for i in range(num_rows):
          row = []
          for j in range(num_cols):
              y1, y2 = i * cell_h, (i + 1) * cell_h
              x1, x2 = j * cell_w, (j + 1) * cell_w

              cell = img_np[y1:y2, x1:x2]
              mean_rgb = cell.reshape(-1, 3).mean(axis=0)
              brightness = int(mean_rgb.mean())
              brightness_8 = round_to_multiple_of_8(brightness)

              row.append(brightness_8)

              # プレビューに描画
              gray = (brightness_8, brightness_8, brightness_8)
              draw.rectangle([x1, y1, x2, y2], fill=gray)

          data.append(row)

      # CSVに保存
      df = pd.DataFrame(data)
      df.to_csv(output_csv, index=False, header=False, encoding="utf-8")
      print(f"CSVを保存しました: {output_csv}")

      # プレビュー保存
      preview.save(output_img)
      print(f"IMGを保存しました: {output_img}")


  # 実行
  image_to_fixed_grid()
  return output_csv

num_rows = 50
num_cols = 80

# CSVファイル名を定義
csv_1 = mosaic("IMG_1091.jpeg", "nina", num_rows, num_cols)
csv_2 = mosaic("IMG_1091.jpeg", "nina", num_rows, num_cols)
csv_3 = mosaic("IMG_1091.jpeg", "nina", num_rows, num_cols)
csv_4 = mosaic("IMG_1091.jpeg", "nina", num_rows, num_cols)
csv_5 = mosaic("IMG_1091.jpeg", "nina", num_rows, num_cols)
csv_6 = mosaic("IMG_1091.jpeg", "nina", num_rows, num_cols)
csv_7 = mosaic("IMG_1091.jpeg", "nina", num_rows, num_cols)

csv_files = [csv_1, csv_2, csv_3, csv_4, csv_5, csv_6, csv_7]

data_frames = [pd.read_csv(f, header=None, nrows=num_rows, usecols=range(num_cols), dtype=int) for f in csv_files]

# Excel作成
wb = Workbook()
ws = wb.active
ws.title = "Sheet"

# 条件付き書式の適用範囲
range_str = "B2:" + chr((num_cols + 1) % 26 + ord('@')) + chr((num_cols + 1) // 26 + ord('@')) + str(num_rows + 1)

print(range_str)

# 16進数8刻みで条件付き書式を作成
hex_values = list(range(0, 256, 8))
if 255 not in hex_values:
    hex_values.append(255)
for val in hex_values:
    hex_str = f"{val:02X}"
    # ARGB形式にする（先頭にFF = fully opaque）
    color_code = f"FF{hex_str}{hex_str}{hex_str}"

    fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")

    # フォント色も同じくARGBで指定
    font_color = f"FF{hex_str}{hex_str}{hex_str}"
    font = Font(color=font_color)

    rule = CellIsRule(operator='equal', formula=[str(val)], fill=fill, font=font)
    ws.conditional_formatting.add(range_str, rule)

# データ範囲に数式を入力
# B2:CC41 → (row=2..41, col=2..81)
for r in range(num_rows):  # 0..39
    for c in range(num_cols):  # 0..79
        vals = []
        for df in data_frames:
            cell_val = df.iat[r, c]
            vals.append(cell_val)
        # CHOOSE関数の数式を作成
        formula = f'=CHOOSE($CN$53,{",".join(map(str, vals))})'
        ws.cell(row=r + 2, column=c + 2, value=formula)
ws.cell(row=53, column=92, value=1)
# 保存
wb.save("EXCEL_output.xlsx")
print("🆗")
